#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Точечное восстановление только по Excel инцидента 9:40:58."""
import json
import re
import sys
import urllib.request
from copy import deepcopy
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
URL = re.search(
    r'url:\s*"([^"]+)"',
    (ROOT / "js" / "gas-config.js").read_text(encoding="utf-8"),
).group(1)
EXCEL = Path(r"c:\Users\Данила\Downloads\инцидент 9-40-58.xlsx")

sys.path.insert(0, str(ROOT / "tools"))
from rollback_burst import apply_field, fmt_field, fetch, post, norm  # noqa: E402


def load_excel_plan():
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    plan = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[2]:
            continue
        plan.append({
            "dealId": str(row[2]).strip(),
            "label": str(row[6] or "").strip(),
            "bilo": row[7],
            "stalo": row[8],
        })
    return plan


def main():
    apply = "--apply" in sys.argv
    plan = load_excel_plan()
    print(f"Excel plan rows: {len(plan)}")
    deal_ids = sorted({p["dealId"] for p in plan})
    print(f"Unique deals: {len(deal_ids)}")

    state = fetch("?action=get")["state"]
    deals_by_id = {d["id"]: deepcopy(d) for d in state.get("deals", []) if d.get("id")}
    print(f"Server deals: {len(deals_by_id)}")

    changes = []
    skipped = 0
    insane = []
    for p in plan:
        deal = deals_by_id.get(p["dealId"])
        if not deal:
            skipped += 1
            continue
        before = fmt_field(deal, p["label"])
        apply_field(deal, p["label"], p["bilo"])
        after = fmt_field(deal, p["label"])
        if norm(before) != norm(after):
            changes.append({**p, "before": str(before)[:80], "after": str(after)[:80]})
        if p["label"] in ("% требований проекта", "% требований пилота", "Вероятность"):
            tr = deal.get("techResearch") or {}
            for key, val in [("productRequirementsPct", tr.get("productRequirementsPct")),
                             ("pilotRequirementsPct", tr.get("pilotRequirementsPct")),
                             ("manualProb", deal.get("manualProb"))]:
                if val is not None and (val > 100 or val < 0):
                    insane.append((p["dealId"], key, val, p["label"]))

    print(f"Fields to change: {len(changes)} (skipped missing deals: {skipped})")
    if insane:
        print("INSANE values after apply:")
        for x in insane[:20]:
            print(" ", x)

  # scan all deals for insane pct
    bad = []
    for d in deals_by_id.values():
        tr = d.get("techResearch") or {}
        for k in ("productRequirementsPct", "pilotRequirementsPct"):
            v = tr.get(k)
            if v is not None and (not isinstance(v, (int, float)) or v > 100 or v < 0):
                bad.append((d["id"], k, v))
        mp = d.get("manualProb")
        if mp is not None and (not isinstance(mp, (int, float)) or mp > 1 or mp < 0):
            bad.append((d["id"], "manualProb", mp))
    print(f"Bad pct/prob on server after plan: {len(bad)}")
    for x in bad[:15]:
        print(" ", x)

    out = ROOT / "tools" / "incident_0940_selective_plan.json"
    out.write_text(json.dumps(changes[:100], ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"Sample changes saved: {out}")

    if not apply:
        print("Preview. Run with --apply")
        return

    recovered = deepcopy(state)
    recovered["deals"] = list(deals_by_id.values())
    res = post({
        "action": "save",
        "state": recovered,
        "forceFull": True,
        "savedBy": "recover-incident-0940-excel",
        "allowMaintenance": True,
    })
    if res.get("error"):
        print("ERROR:", res["error"])
        sys.exit(1)
    print("Applied:", res.get("updatedAt"), "auditRows=", res.get("auditRows"))


if __name__ == "__main__":
    main()
